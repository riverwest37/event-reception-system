import base64
import re
import subprocess
from datetime import datetime
from email.header import decode_header
from pathlib import Path


from openpyxl import Workbook, load_workbook
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build




SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]


BASE_DIR = Path.home() / "gmail_sorter"
EXCEL_PATH = BASE_DIR / "受付台帳.xlsx"
TOKEN_PATH = BASE_DIR / "token.json"


# Dropbox上の正式台帳
DROPBOX_EXCEL_PATH = "dropbox:受付台帳.xlsx"
DROPBOX_REMOTE_FOLDER = "dropbox:"




# ============================================================
# オンライン交流会の受付期間
# お問い合わせ・会員申込には、この期間制限はかけない
# ============================================================


RECEPTION_START_DATE = "2026/7/10"
RECEPTION_END_DATE = "2026/8/5"




# Jimdo通知メール全体を対象にする
GMAIL_BASE_QUERY = 'from:no-reply@jimdo.jp'




SHEET_ONLINE = "オンライン交流会"
SHEET_CONTACT = "お問い合わせ"
SHEET_MEMBER = "会員申込"




HEADERS_ONLINE = [
    "受付No",
    "受付日時",
    "氏名",
    "メールアドレス",
    "電話番号",
    "現住所",
    "出身",
    "会員区分",
    "8/2練習会",
    "8/4練習会",
    "8/5練習会",
    "交流会8/5",
    "会側確認状況",
    "委託先確認状況",
    "全体対応状況",
    "メッセージ",
    "メール件名",
    "GmailメッセージID",
]




HEADERS_CONTACT = [
    "受付No",
    "受付日時",
    "氏名",
    "メールアドレス",
    "問い合わせ内容",
    "対応担当",
    "対応状況",
    "備考",
    "メール件名",
    "GmailメッセージID",
]




HEADERS_MEMBER = [
    "受付No",
    "受付日時",
    "氏名",
    "メールアドレス",
    "申込区分",
    "現住所",
    "電話番号",
    "出身地",
    "性別",
    "年齢",
    "会員管理確認",
    "対応状況",
    "備考",
    "メール件名",
    "GmailメッセージID",
]




def build_search_query():
    """
    Jimdo通知メールを取得する。
    ここでは期間制限をかけない。
    オンライン交流会だけ、後でPython側で期間判定する。
    """
    return GMAIL_BASE_QUERY




SEARCH_QUERY = build_search_query()




def parse_date_ymd(text):
    return datetime.strptime(text, "%Y/%m/%d").date()




def online_reception_range():
    start = parse_date_ymd(RECEPTION_START_DATE)
    end = parse_date_ymd(RECEPTION_END_DATE)
    return start, end




def gmail_internal_date_to_datetime(message):
    """
    Gmail APIの internalDate はミリ秒。
    """
    internal_date = message.get("internalDate")
    if not internal_date:
        return datetime.now()


    return datetime.fromtimestamp(int(internal_date) / 1000)




def is_online_reception_period(message_dt):
    start, end = online_reception_range()
    target_date = message_dt.date()
    return start <= target_date <= end




def decode_mime_header(value):
    if not value:
        return ""


    result = ""
    for part, encoding in decode_header(value):
        if isinstance(part, bytes):
            result += part.decode(encoding or "utf-8", errors="replace")
        else:
            result += part
    return result




def get_gmail_service():
    """
    既存の token.json を使ってGmailを読み取る。
    Gmail APIは readonly のまま。
    """
    if not TOKEN_PATH.exists():
        raise FileNotFoundError(f"token.json が見つかりません: {TOKEN_PATH}")


    creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)


    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())


    if not creds or not creds.valid:
        raise RuntimeError("Gmail認証情報が無効です。token.jsonを確認してください。")


    return build("gmail", "v1", credentials=creds)




def get_header(headers, name):
    for h in headers:
        if h.get("name", "").lower() == name.lower():
            return decode_mime_header(h.get("value", ""))
    return ""




def decode_body_data(data):
    if not data:
        return ""


    try:
        decoded = base64.urlsafe_b64decode(data.encode("utf-8"))
        return decoded.decode("utf-8", errors="replace")
    except Exception:
        return ""




def extract_text_from_payload(payload):
    """
    Gmail payload から本文テキストを取り出す。
    text/plain を優先し、なければ text/html も簡易的に文字として読む。
    """
    if not payload:
        return ""


    mime_type = payload.get("mimeType", "")
    body = payload.get("body", {})
    data = body.get("data")


    if data and mime_type in ("text/plain", "text/html"):
        return decode_body_data(data)


    parts = payload.get("parts", [])
    texts = []


    # text/plainを優先
    for part in parts:
        if part.get("mimeType") == "text/plain":
            text = extract_text_from_payload(part)
            if text:
                texts.append(text)


    if texts:
        return "\n".join(texts)


    # text/plainがない場合は全partを再帰的に読む
    for part in parts:
        text = extract_text_from_payload(part)
        if text:
            texts.append(text)


    return "\n".join(texts)




def normalize_text(text):
    if not text:
        return ""
    return text.replace("\r\n", "\n").replace("\r", "\n")




def extract_value(body, label):
    """
    「項目名: 値」の形式から値を取り出す。
    次の項目行に行くまでを対象にする。
    """
    body = normalize_text(body)
    pattern = re.escape(label) + r"\s*(.*)"
    match = re.search(pattern, body)
    if not match:
        return ""


    value = match.group(1).strip()


    # 値が次行以降に続く場合の簡易対応
    lines = body[match.end():].splitlines()
    for line in lines:
        line = line.strip()
        if not line:
            break
        if re.search(r".+[:：]$", line):
            break
        if re.search(r".+[:：]\s*.+", line):
            break
        value += " " + line


    return value.strip()




def extract_by_patterns(body, patterns):
    body = normalize_text(body)


    for pattern in patterns:
        match = re.search(pattern, body)
        if match:
            return match.group(1).strip()


    return ""




def detect_mail_type(subject, body):
    """
    Jimdo通知メールの件名・本文・URLから申請種別を判定する。
    """
    text = f"{subject}\n{body}"


    # オンライン交流会
    if "オンライン交流会" in text and "申込書" in text:
        return SHEET_ONLINE


    # 会員申込
    if "会員募集/会員募集申込用紙" in text:
        return SHEET_MEMBER


    if "会員申込区分" in text:
        return SHEET_MEMBER


    if "東京四万十会申込書" in text:
        return SHEET_MEMBER


    # お問い合わせ
    if "お問い合わせ" in text or "お問合わせ" in text:
        return SHEET_CONTACT


    if "東京四万十会HPより問合わせ" in text:
        return SHEET_CONTACT


    return ""




def get_or_create_sheet(wb, sheet_name, headers):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)


    # ヘッダーだけを整える。既存データ行は触らない。
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header


    return ws




def prepare_workbook():
    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
    else:
        wb = Workbook()
        default_ws = wb.active
        default_ws.title = SHEET_ONLINE


    ws_online = get_or_create_sheet(wb, SHEET_ONLINE, HEADERS_ONLINE)
    ws_contact = get_or_create_sheet(wb, SHEET_CONTACT, HEADERS_CONTACT)
    ws_member = get_or_create_sheet(wb, SHEET_MEMBER, HEADERS_MEMBER)


    return wb, {
        SHEET_ONLINE: ws_online,
        SHEET_CONTACT: ws_contact,
        SHEET_MEMBER: ws_member,
    }




def get_message_id_column(sheet_name):
    if sheet_name == SHEET_ONLINE:
        return 18
    if sheet_name == SHEET_CONTACT:
        return 10
    if sheet_name == SHEET_MEMBER:
        return 15
    return None




def get_registered_message_ids(ws, sheet_name):
    id_col = get_message_id_column(sheet_name)
    ids = set()


    if not id_col or ws.max_row < 2:
        return ids


    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=id_col).value
        if value:
            ids.add(str(value))


    return ids




def get_all_registered_ids(sheets):
    """
    3シート全体で重複防止する。
    同じGmailメッセージIDは、どのシートにも二重登録しない。
    """
    all_ids = set()


    for sheet_name, ws in sheets.items():
        all_ids |= get_registered_message_ids(ws, sheet_name)


    return all_ids




def next_receipt_no(ws):
    if ws.max_row < 2:
        return "001"


    last_no = ws.cell(row=ws.max_row, column=1).value


    try:
        return str(int(last_no) + 1).zfill(3)
    except Exception:
        return str(ws.max_row).zfill(3)




def extract_online_row(ws, body, subject, msg_id):
    name = extract_value(body, "お名前:")
    email = extract_value(body, "メールアドレス:")
    address = extract_value(body, "現住所【県・都市のみ】:")
    phone = extract_value(body, "電話番号:")
    origin = extract_value(body, "ご出身:")
    desired = extract_value(body, "希望日「複数可」:")
    member = extract_value(body, "東京四万十会会員:")


    message_text = extract_by_patterns(body, [
        r"メッセージ「オンライン交流会申込書」\s*:\s*(.*)",
        r"メッセージ\s*「オンライン交流会申込書」\s*:\s*(.*)",
        r"メッセージ.*オンライン交流会申込書.*:\s*(.*)",
    ])


    receipt_no = next_receipt_no(ws)


    return [
        receipt_no,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        name,
        email,
        phone,
        address,
        origin,
        member,
        "○" if "練習会8/2" in desired else "",
        "○" if "練習会8/4" in desired else "",
        "○" if "練習会8/5" in desired else "",
        "○" if "交流会8/5" in desired else "",
        "未確認",
        "未確認",
        "未対応",
        message_text,
        subject,
        msg_id,
    ]




def extract_contact_row(ws, body, subject, msg_id):
    name = extract_value(body, "お名前:")
    email = extract_value(body, "メールアドレス:")


    inquiry = extract_by_patterns(body, [
        r"東京四万十会HPより問合わせ\s*:\s*(.*)",
        r"東京四万十会HPより問い合わせ\s*:\s*(.*)",
        r"東京四万十会HPよりお問合わせ\s*:\s*(.*)",
    ])


    receipt_no = next_receipt_no(ws)


    return [
        receipt_no,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        name,
        email,
        inquiry,
        "",
        "未対応",
        "",
        subject,
        msg_id,
    ]




def extract_member_row(ws, body, subject, msg_id):
    name = extract_value(body, "名前:")
    if not name:
        name = extract_value(body, "お名前:")


    email = extract_value(body, "メールアドレス:")


    apply_type = extract_by_patterns(body, [
        r"会員申込区分【東京四万十会申込書】\s*:\s*(.*)",
        r"会員申込区分.*東京四万十会申込書.*:\s*(.*)",
        r"会員申込区分\s*:\s*(.*)",
    ])


    address = extract_value(body, "現住所:")
    phone = extract_value(body, "電話:")


    origin = extract_by_patterns(body, [
        r"出身地（四万十町の方は地区名、以外の方は県・都市名）\s*:\s*(.*)",
        r"出身地.*:\s*(.*)",
    ])


    gender = extract_value(body, "性別:")
    age = extract_value(body, "年齢:")


    receipt_no = next_receipt_no(ws)


    return [
        receipt_no,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        name,
        email,
        apply_type,
        address,
        phone,
        origin,
        gender,
        age,
        "未確認",
        "未対応",
        "",
        subject,
        msg_id,
    ]




def copy_from_dropbox():
    """
    Dropbox上の最新版 受付台帳.xlsx をPi3へ取り込む。


    重要：
    Dropbox上で担当者が手入力した
    「会側確認状況」「委託先確認状況」「全体対応状況」
    などの更新内容を消さないため、台帳を開く前に必ず実行する。
    """
    try:
        subprocess.run(
            ["rclone", "copyto", DROPBOX_EXCEL_PATH, str(EXCEL_PATH)],
            check=True
        )
        print("Dropboxから受付台帳.xlsxを取得しました。")
    except subprocess.CalledProcessError as e:
        print("Dropboxからの取得に失敗しました。Pi3側の台帳を使用します。")
        print(e)




def copy_to_dropbox():
    if not EXCEL_PATH.exists():
        print(f"Dropboxコピー対象がありません: {EXCEL_PATH}")
        return


    subprocess.run(
        ["rclone", "copy", str(EXCEL_PATH), DROPBOX_REMOTE_FOLDER],
        check=True
    )


    print("Dropboxへ受付台帳.xlsxをコピーしました。")




def main():
    # 先にDropbox上の最新版をPi3へ取り込む。
    # これによりDropboxで手入力した確認状況を保持する。
    copy_from_dropbox()


    service = get_gmail_service()
    wb, sheets = prepare_workbook()


    registered_ids = get_all_registered_ids(sheets)


    result = service.users().messages().list(
        userId="me",
        q=SEARCH_QUERY,
        maxResults=100
    ).execute()


    messages = result.get("messages", [])


    print(f"Jimdo検索条件: {SEARCH_QUERY}")
    print(f"オンライン交流会受付期間: {RECEPTION_START_DATE} 〜 {RECEPTION_END_DATE}")
    print(f"取得メール数: {len(messages)}")
    print("-" * 60)


    added_online = 0
    added_contact = 0
    added_member = 0
    skipped_registered = 0
    skipped_period = 0
    skipped_unknown = 0


    for item in messages:
        msg_id = item.get("id")


        if not msg_id:
            continue


        if msg_id in registered_ids:
            print(f"登録済みのためスキップ: {msg_id}")
            skipped_registered += 1
            continue


        message = service.users().messages().get(
            userId="me",
            id=msg_id,
            format="full"
        ).execute()


        payload = message.get("payload", {})
        headers = payload.get("headers", [])


        subject = get_header(headers, "Subject")
        body = extract_text_from_payload(payload)
        message_dt = gmail_internal_date_to_datetime(message)


        sheet_name = detect_mail_type(subject, body)


        if not sheet_name:
            print(f"分類不能のためスキップ: {msg_id} / {subject}")
            skipped_unknown += 1
            continue


        if sheet_name == SHEET_ONLINE:
            if not is_online_reception_period(message_dt):
                print(f"期間外のためスキップ: {msg_id} / {subject}")
                skipped_period += 1
                continue


            ws = sheets[SHEET_ONLINE]
            row = extract_online_row(ws, body, subject, msg_id)
            ws.append(row)
            registered_ids.add(msg_id)
            added_online += 1
            print(f"追加[オンライン交流会]: {row[0]} {row[2]}")


        elif sheet_name == SHEET_CONTACT:
            ws = sheets[SHEET_CONTACT]
            row = extract_contact_row(ws, body, subject, msg_id)
            ws.append(row)
            registered_ids.add(msg_id)
            added_contact += 1
            print(f"追加[お問い合わせ]: {row[0]} {row[2]}")


        elif sheet_name == SHEET_MEMBER:
            ws = sheets[SHEET_MEMBER]
            row = extract_member_row(ws, body, subject, msg_id)
            ws.append(row)
            registered_ids.add(msg_id)
            added_member += 1
            print(f"追加[会員申込]: {row[0]} {row[2]}")


    wb.save(EXCEL_PATH)


    print("-" * 60)
    print(f"追加[オンライン交流会]: {added_online}")
    print(f"追加[お問い合わせ]: {added_contact}")
    print(f"追加[会員申込]: {added_member}")
    print(f"登録済みスキップ: {skipped_registered}")
    print(f"期間外スキップ: {skipped_period}")
    print(f"分類不能スキップ: {skipped_unknown}")
    print(f"台帳保存先: {EXCEL_PATH}")


    copy_to_dropbox()




if __name__ == "__main__":
    main()
