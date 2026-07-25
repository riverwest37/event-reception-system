from pathlib import Path
from openpyxl import load_workbook


BASE_DIR = Path.home() / "gmail_sorter"
EXCEL_PATH = BASE_DIR / "受付台帳.xlsx"


SHEET_ONLINE = "オンライン交流会"
SHEET_CONTACT = "お問い合わせ"
SHEET_MEMBER = "会員申込"




def short(value, width=18):
    """長い文字を見やすく短縮する"""
    if value is None:
        return ""


    text = str(value)


    if len(text) > width:
        return text[:width - 1] + "…"


    return text




def print_table(ws, columns, widths, title):
    print()
    print("=" * 180)
    print(f"【{title}】")
    print(f"登録件数: {max(ws.max_row - 1, 0)} 件")
    print("=" * 180)


    if ws.max_row < 1:
        print("見出し行がありません。")
        return


    # 見出し
    headers = []
    for col in columns:
        width = widths.get(col, 14)
        headers.append(short(ws.cell(row=1, column=col).value, width))


    print(" | ".join(
        f"{headers[i]:<{widths.get(columns[i], 14)}}"
        for i in range(len(columns))
    ))


    print("-" * 180)


    # データ行
    if ws.max_row < 2:
        print("データはありません。")
        return


    for row in range(2, ws.max_row + 1):
        values = []


        for col in columns:
            width = widths.get(col, 14)
            values.append(short(ws.cell(row=row, column=col).value, width))


        print(" | ".join(
            f"{values[i]:<{widths.get(columns[i], 14)}}"
            for i in range(len(columns))
        ))




def show_online(ws):
    columns = [
        1,   # 受付No
        2,   # 受付日時
        3,   # 氏名
        4,   # メールアドレス
        5,   # 電話番号
        8,   # 会員区分
        9,   # 8/2練習会
        10,  # 8/4練習会
        11,  # 8/5練習会
        12,  # 交流会8/5
        13,  # 会員管理確認
        14,  # 運用担当確認
        15,  # 対応状況
        16,  # メッセージ
    ]


    widths = {
        1: 8,
        2: 18,
        3: 14,
        4: 24,
        5: 14,
        8: 10,
        9: 10,
        10: 10,
        11: 10,
        12: 10,
        13: 12,
        14: 12,
        15: 10,
        16: 30,
    }


    print_table(ws, columns, widths, SHEET_ONLINE)


    # メッセージ詳細
    print()
    print("メッセージ詳細")
    print("-" * 80)


    has_message = False


    for row in range(2, ws.max_row + 1):
        receipt_no = ws.cell(row=row, column=1).value
        name = ws.cell(row=row, column=3).value
        message = ws.cell(row=row, column=16).value


        if message:
            has_message = True
            print(f"{receipt_no} {name}: {message}")


    if not has_message:
        print("メッセージはありません。")


    print("-" * 80)




def show_contact(ws):
    columns = [
        1,   # 受付No
        2,   # 受付日時
        3,   # 氏名
        4,   # メールアドレス
        5,   # 問い合わせ内容
        6,   # 対応担当
        7,   # 対応状況
        8,   # 備考
    ]


    widths = {
        1: 8,
        2: 18,
        3: 14,
        4: 24,
        5: 40,
        6: 12,
        7: 10,
        8: 20,
    }


    print_table(ws, columns, widths, SHEET_CONTACT)


    # 問い合わせ詳細
    print()
    print("問い合わせ内容詳細")
    print("-" * 80)


    has_inquiry = False


    for row in range(2, ws.max_row + 1):
        receipt_no = ws.cell(row=row, column=1).value
        name = ws.cell(row=row, column=3).value
        inquiry = ws.cell(row=row, column=5).value


        if inquiry:
            has_inquiry = True
            print(f"{receipt_no} {name}: {inquiry}")


    if not has_inquiry:
        print("問い合わせ内容はありません。")


    print("-" * 80)




def show_member(ws):
    columns = [
        1,   # 受付No
        2,   # 受付日時
        3,   # 氏名
        4,   # メールアドレス
        5,   # 申込区分
        6,   # 現住所
        7,   # 電話番号
        8,   # 出身地
        9,   # 性別
        10,  # 年齢
        11,  # 会員管理確認
        12,  # 対応状況
        13,  # 備考
    ]


    widths = {
        1: 8,
        2: 18,
        3: 14,
        4: 24,
        5: 28,
        6: 24,
        7: 14,
        8: 16,
        9: 6,
        10: 10,
        11: 12,
        12: 10,
        13: 20,
    }


    print_table(ws, columns, widths, SHEET_MEMBER)


    # 会員申込詳細
    print()
    print("会員申込詳細")
    print("-" * 80)


    if ws.max_row < 2:
        print("会員申込データはありません。")
        print("-" * 80)
        return


    for row in range(2, ws.max_row + 1):
        receipt_no = ws.cell(row=row, column=1).value
        name = ws.cell(row=row, column=3).value
        apply_type = ws.cell(row=row, column=5).value
        address = ws.cell(row=row, column=6).value
        origin = ws.cell(row=row, column=8).value
        gender = ws.cell(row=row, column=9).value
        age = ws.cell(row=row, column=10).value


        print(
            f"{receipt_no} {name}: "
            f"区分={apply_type}, "
            f"住所={address}, "
            f"出身={origin}, "
            f"性別={gender}, "
            f"年齢={age}"
        )


    print("-" * 80)




def show_summary(wb):
    print()
    print("受付台帳サマリー")
    print("=" * 60)


    for sheet_name in [SHEET_ONLINE, SHEET_CONTACT, SHEET_MEMBER]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            count = max(ws.max_row - 1, 0)
            print(f"{sheet_name}: {count} 件")
        else:
            print(f"{sheet_name}: シートなし")


    print("=" * 60)




def main():
    if not EXCEL_PATH.exists():
        print(f"受付台帳が見つかりません: {EXCEL_PATH}")
        return


    wb = load_workbook(EXCEL_PATH, data_only=True)


    print(f"受付台帳: {EXCEL_PATH}")
    print(f"シート一覧: {', '.join(wb.sheetnames)}")


    show_summary(wb)


    if SHEET_ONLINE in wb.sheetnames:
        show_online(wb[SHEET_ONLINE])
    else:
        print(f"シートがありません: {SHEET_ONLINE}")


    if SHEET_CONTACT in wb.sheetnames:
        show_contact(wb[SHEET_CONTACT])
    else:
        print(f"シートがありません: {SHEET_CONTACT}")


    if SHEET_MEMBER in wb.sheetnames:
        show_member(wb[SHEET_MEMBER])
    else:
        print(f"シートがありません: {SHEET_MEMBER}")




if __name__ == "__main__":
    main()
