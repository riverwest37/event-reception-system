from pathlib import Path
from datetime import datetime, date
import smtplib
from email.mime.text import MIMEText
from openpyxl import load_workbook


from mail_config import SMTP_USER, SMTP_APP_PASSWORD, MAIL_TO, MAIL_CC




BASE_DIR = Path.home() / "gmail_sorter"
EXCEL_PATH = BASE_DIR / "受付台帳.xlsx"


SHEET_ONLINE = "オンライン交流会"




def value(row, index):
    """
    openpyxlの行データから安全に値を取り出す。
    indexは0始まり。
    """
    if index >= len(row):
        return ""
    return row[index] if row[index] is not None else ""




def make_daily_report():
    now = datetime.now()


    if not EXCEL_PATH.exists():
        return (
            "【東京四万十会】オンライン交流会 Daily Report",
            f"""東京四万十会 オンライン交流会 Daily Report


発行日時: {now.strftime('%Y-%m-%d %H:%M')}


受付台帳が見つかりません。


確認対象:
{EXCEL_PATH}
"""
        )


    wb = load_workbook(EXCEL_PATH, data_only=True)


    if SHEET_ONLINE not in wb.sheetnames:
        return (
            "【東京四万十会】オンライン交流会 Daily Report",
            f"""東京四万十会 オンライン交流会 Daily Report


発行日時: {now.strftime('%Y-%m-%d %H:%M')}


「オンライン交流会」シートが見つかりません。
"""
        )


    ws = wb[SHEET_ONLINE]


    rows = list(ws.iter_rows(min_row=2, values_only=True))


    total_count = 0
    today_count = 0
    unchecked_member = 0
    unchecked_operator = 0
    not_done = 0


    today_rows = []


    today_str = date.today().strftime("%Y-%m-%d")


    for row in rows:
        receipt_no = value(row, 0)
        if not receipt_no:
            continue


        total_count += 1


        received_at = str(value(row, 1))
        name = value(row, 2)
        email = value(row, 3)
        phone = value(row, 4)
        member_type = value(row, 7)


        r_82 = value(row, 8)
        r_84 = value(row, 9)
        r_85 = value(row, 10)
        r_event85 = value(row, 11)


        member_check = value(row, 12)
        operator_check = value(row, 13)
        status = value(row, 14)
        message = value(row, 15)


        if str(member_check) == "未確認":
            unchecked_member += 1


        if str(operator_check) == "未確認":
            unchecked_operator += 1


        if str(status) == "未対応":
            not_done += 1


        if received_at.startswith(today_str):
            today_count += 1
            today_rows.append({
                "受付No": receipt_no,
                "受付日時": received_at,
                "氏名": name,
                "メール": email,
                "電話": phone,
                "会員区分": member_type,
                "8/2練習会": r_82,
                "8/4練習会": r_84,
                "8/5練習会": r_85,
                "交流会8/5": r_event85,
                "メッセージ": message,
            })


    subject = f"【東京四万十会】オンライン交流会 Daily Report {now.strftime('%Y-%m-%d')}"


    lines = []
    lines.append("東京四万十会 オンライン交流会 Daily Report")
    lines.append("")
    lines.append(f"発行日時: {now.strftime('%Y-%m-%d %H:%M')}")
    lines.append("")
    lines.append("【本日の新規申込】")
    lines.append(f"{today_count}件")
    lines.append("")


    if today_rows:
        for item in today_rows:
            lines.append(f"受付No.{item['受付No']}")
            lines.append(f"氏名: {item['氏名']}")
            lines.append(f"会員区分: {item['会員区分']}")
            lines.append(f"メール: {item['メール']}")
            lines.append(f"電話: {item['電話']}")
            lines.append(
                "希望日: "
                f"8/2練習会={item['8/2練習会']} "
                f"8/4練習会={item['8/4練習会']} "
                f"8/5練習会={item['8/5練習会']} "
                f"交流会8/5={item['交流会8/5']}"
            )
            lines.append(f"メッセージ: {item['メッセージ']}")
            lines.append("-" * 40)
    else:
        lines.append("本日の新規申込はありません。")
        lines.append("")


    lines.append("")
    lines.append("【累計】")
    lines.append(f"オンライン交流会申込: {total_count}件")
    lines.append("")
    lines.append("【確認状況】")
    lines.append(f"会員管理確認 未確認: {unchecked_member}件")
    lines.append(f"運用担当確認 未確認: {unchecked_operator}件")
    lines.append(f"対応状況 未対応: {not_done}件")
    lines.append("")
    lines.append("受付台帳はDropboxの「受付台帳.xlsx」を確認してください。")
    lines.append("")
    lines.append("このメールはPi3の自動受付システムから送信されています。")


    return subject, "\n".join(lines)




def send_mail(subject, body):
    msg = MIMEText(body, "plain", "utf-8")
    msg["Subject"] = subject
    msg["From"] = SMTP_USER
    msg["To"] = MAIL_TO


    recipients = [x.strip() for x in MAIL_TO.split(",") if x.strip()]


    if MAIL_CC:
        msg["Cc"] = MAIL_CC
        recipients += [x.strip() for x in MAIL_CC.split(",") if x.strip()]


    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(SMTP_USER, SMTP_APP_PASSWORD)
        server.send_message(msg, from_addr=SMTP_USER, to_addrs=recipients)




def main():
    subject, body = make_daily_report()
    send_mail(subject, body)
    print("Daily Reportを送信しました。")
    print(f"送信先: {MAIL_TO}")




if __name__ == "__main__":
    main()
